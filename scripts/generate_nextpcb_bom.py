#!/usr/bin/env python3
"""Convert JLCPCB BOM (production/bom.csv) to NextPCB Excel format.

Reads the KiCad Fabrication Toolkit BOM and outputs a NextPCB-compatible
Excel file with MPN-based part identification.

Usage: python3 scripts/generate_nextpcb_bom.py
Output: production/nextpcb_bom.xlsx
"""

import csv
import os
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("Error: openpyxl required. Install with: pip3 install openpyxl")
    raise SystemExit(1)

PROJECT_DIR = Path(__file__).resolve().parent.parent
BOM_CSV = PROJECT_DIR / "production" / "bom.csv"
OUTPUT_XLS = PROJECT_DIR / "production" / "nextpcb_bom.xlsx"

# LCSC part number -> (MPN, Manufacturer, Description) mapping
# Active components have MPNs from schematic; passives looked up from LCSC
LCSC_TO_MPN = {
    # Capacitors
    "C105226":   ("CL05A226MQ5QUNC",     "Samsung Electro-Mechanics", "22uF 6.3V X5R 0402"),
    "C440198":   ("GRM21BR61H106KE43L",   "Murata",                   "10uF 50V X5R 0805"),
    "C307331":   ("CL05B104KB54PNC",      "Samsung Electro-Mechanics", "100nF 50V X7R 0402"),
    "C2762594":  ("CL10A226MO7JZNC",      "Samsung Electro-Mechanics", "22uF 6.3V X5R 0603"),
    "C181043":   ("GRM033R6YA104KE14D",   "Murata",                   "100nF 16V X5R 0201"),
    # Ferrite beads
    "C6132061":  ("BLM03PX121SN1D",       "Murata",                   "Ferrite bead 120R 900mA 0201"),
    # Resistors
    "C473542":   ("0201WMF150JTEE",       "UNI-ROYAL",               "15R 1% 0201"),
    "C473048":   ("0201WMF1002TEE",       "UNI-ROYAL",               "10k 1% 0201"),
    "C270365":   ("0201WMF1001TEE",       "UNI-ROYAL",               "1k 1% 0201"),
    "C423759":   ("0201WMF1243TEE",       "UNI-ROYAL",               "124k 1% 0201"),
    "C473535":   ("0201WMJ0221TEE",       "UNI-ROYAL",               "220R 5% 0201"),
    "C695806":   ("ASR-S-3-0.2F",         "Yezhan",                   "0.2mOhm 3W 2512 shunt"),
    # Active - ICs
    "C59135":    ("INA199A1DCKR",         "Texas Instruments",        "Current sense amp SC-70-6"),
    "C2765098":  ("AT32F421G8U7",         "Artery",                   "MCU ARM Cortex-M4 QFN-28"),
    "C41414478": ("NSG2065Q",             "Novosense",                "3-phase gate driver QFN-24"),
    "C5383002":  ("LMR51420YDDCR",        "Texas Instruments",        "Buck converter 4.2-36V SOT-23-6"),
    "C524780":   ("TLV76733DRVR",          "Texas Instruments",        "3.3V LDO 150mA WSON-6"),
    "C22466709": ("SP40N03GNJ",           "JSC",                      "N-MOSFET 40V/75A DFN-8 3x3mm"),
    # Passive - special
    "C48391583": ("XRIM160808SR47MBCD",   "XR",                       "470nH inductor 0603"),
    "C2845367":  ("HC-1.0-8PWT",          "HCTL",                     "8-pin 1mm pitch SMD connector"),
}

# Parts to exclude from assembly (no LCSC part or not populated)
EXCLUDE_DESIGNATORS = {"TP1", "TP2", "TP3", "TP4", "TP5", "TP6", "TP7", "TP8"}


def parse_bom_csv():
    """Parse the JLCPCB-format BOM CSV."""
    rows = []
    with open(BOM_CSV, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            designators = [d.strip() for d in row["Designator"].split(",")]
            lcsc = row.get("LCSC Part #", "").strip()
            value = row.get("Value", "").strip()
            footprint = row.get("Footprint", "").strip()
            quantity = int(row.get("Quantity", len(designators)))
            rows.append({
                "designators": designators,
                "designator_str": row["Designator"].strip(),
                "footprint": footprint,
                "quantity": quantity,
                "value": value,
                "lcsc": lcsc,
            })
    return rows


def make_nextpcb_bom():
    """Generate NextPCB Excel BOM."""
    bom_rows = parse_bom_csv()
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"

    # Header style
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = [
        "Designator*",
        "Quantity*",
        "Manufacturer Part Number*",
        "Manufacturer",
        "Package/Footprint",
        "Description",
        "Procurement Type",
        "Customer Note",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    row_num = 2
    skipped = []
    for bom_row in bom_rows:
        # Filter out test points
        desigs = [d for d in bom_row["designators"] if d not in EXCLUDE_DESIGNATORS]
        if not desigs:
            skipped.append(bom_row["designator_str"])
            continue

        lcsc = bom_row["lcsc"]
        value = bom_row["value"]

        # Determine MPN, manufacturer, description
        if lcsc and lcsc in LCSC_TO_MPN:
            mpn, mfr, desc = LCSC_TO_MPN[lcsc]
        elif value and value != "~":
            # Use value as MPN (some active parts have MPN as value)
            mpn = value
            mfr = ""
            desc = ""
        else:
            # No part info - mark as DNP
            mpn = ""
            mfr = ""
            desc = ""

        # Determine procurement type
        procurement = ""
        if not lcsc and (not value or value == "~"):
            procurement = "DNP"

        desig_str = ", ".join(desigs)
        qty = len(desigs)

        ws.cell(row=row_num, column=1, value=desig_str).border = thin_border
        ws.cell(row=row_num, column=2, value=qty).border = thin_border
        ws.cell(row=row_num, column=3, value=mpn).border = thin_border
        ws.cell(row=row_num, column=4, value=mfr).border = thin_border
        ws.cell(row=row_num, column=5, value=bom_row["footprint"]).border = thin_border
        ws.cell(row=row_num, column=6, value=desc).border = thin_border
        ws.cell(row=row_num, column=7, value=procurement).border = thin_border
        ws.cell(row=row_num, column=8, value=f"LCSC: {lcsc}" if lcsc else "").border = thin_border

        row_num += 1

    # Auto-width columns
    for col in range(1, len(headers) + 1):
        max_len = len(headers[col - 1])
        for row in range(2, row_num):
            val = ws.cell(row=row, column=col).value
            if val and len(str(val)) > max_len:
                max_len = len(str(val))
        ws.column_dimensions[chr(64 + col)].width = min(max_len + 2, 60)

    wb.save(OUTPUT_XLS)
    print(f"NextPCB BOM saved to: {OUTPUT_XLS}")
    print(f"Total lines: {row_num - 2}")
    if skipped:
        print(f"Skipped (test points/DNP): {skipped}")


if __name__ == "__main__":
    make_nextpcb_bom()
